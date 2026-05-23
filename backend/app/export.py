"""Excel export for detections over a date range."""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from . import db


def build_xlsx(start: str, end: str) -> bytes:
    wb = Workbook()

    # Sheet 1: detections
    ws = wb.active
    ws.title = "Detections"
    headers = ["ID", "Timestamp", "Date", "Category", "Color", "Hue", "Saturation", "Value"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for row in db.all_in_range(start, end):
        ws.append([
            row["id"], row["ts"], row["day"], row["category"], row["color_name"],
            row["hue"], row["saturation"], row["value"],
        ])

    widths = [8, 22, 14, 10, 12, 8, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Sheet 2: per-day totals
    ws2 = wb.create_sheet("Daily Totals")
    ws2.append(["Date", "A", "B", "C", "D", "E", "Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for d in db.totals_range(start, end):
        ws2.append([d["day"], d["A"], d["B"], d["C"], d["D"], d["E"], d["total"]])

    for i, w in enumerate([14, 8, 8, 8, 8, 8, 10], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(start: str, end: str) -> str:
    if start == end:
        return f"cap-sort-{start}.xlsx"
    return f"cap-sort-{start}_to_{end}.xlsx"
